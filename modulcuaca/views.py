from django.shortcuts import render
from django.utils import timezone
from django.conf import settings
from django.template import RequestContext
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db.models import Avg, Max, Min, Sum
from .models import TabelSuhu, TabelKelembaban, TabelTekananUdara
from datetime import datetime, timedelta
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
import json, string

# api_suhu, api_kelembaban, api_tekanan udara are import data from database and convert to JSON
# Kenapa di convert ke JSON, karena kita sedang membuat API, dan API memliki format data JSON atau XML
# Fungsi JSON ? sebagai format untuk pertukaran data
# JSON berfungsi untuk pertukaran data dari awalnya data yg diambil dengan bahasa python, lalu jadi sebuah API, yang kemudian kita parse data Json tsb dari API untuk digunakan di bahasa javascript dalam membuat grafik

def web_grafik(request):
	return render(request, 'modulcuaca/grafik.html')

def web_raspi(request):
	return render(request, 'modulcuaca/raspi.html')

def web_home(request):
    return render(request, 'modulcuaca/home.html')

def export_suhu_to_xlsx(request):
    temp_queryset = TabelSuhu.objects.filter(api_key='DEFISR98')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}_data-suhu.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Temperature Control'

    # Define the titles for columns
    columns = [
        'Waktu',
        'API_key',
        'Suhu',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for temp in temp_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            temp.time,
            temp.api_key,
            temp.nilai_suhu,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

def export_kelembaban_to_xlsx(request):
    hum_queryset = TabelKelembaban.objects.filter(api_key='DEFISR98')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}_data-kelembaban.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Temperature Control'

    # Define the titles for columns
    columns = [
        'Waktu',
        'API_key',
        'Kelembaban',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for hum in hum_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            hum.time,
            hum.api_key,
            hum.nilai_kelembaban,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

def export_tekananudara_to_xlsx(request):
    press_queryset = TabelTekananUdara.objects.filter(api_key='DEFISR98')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}_data-tekananudara.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Temperature Control'

    # Define the titles for columns
    columns = [
        'Waktu',
        'API_key',
        'Tekanan Udara',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for press in press_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            press.time,
            press.api_key,
            press.nilai_tekananudara,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

def api_temp(request):
    querytemp = TabelSuhu.objects.filter(api_key='DEFISR98')

    listing = []
    for item in querytemp:
	    listing.append({
	            'temp_time'     : item.time,
	            'temp_value' 	: item.nilai_suhu
	        })
    return JsonResponse(listing, safe=False)

def api_summary(request):
    querytemp = TabelSuhu.objects.filter(api_key='DEFISR98')
    queryhum = TabelKelembaban.objects.filter(api_key='DEFISR98')
    querypress = TabelTekananUdara.objects.filter(api_key='DEFISR98')
    datapress = querypress.last()
    datatemp = querytemp.last()
    datahum = queryhum.last()

    min_hum = TabelKelembaban.objects.all().aggregate(Min('nilai_kelembaban'))
    avg_hum = TabelKelembaban.objects.all().aggregate(Avg('nilai_kelembaban'))
    max_hum = TabelKelembaban.objects.all().aggregate(Max('nilai_kelembaban'))

    min_temp = TabelSuhu.objects.all().aggregate(Min('nilai_suhu'))
    avg_temp = TabelSuhu.objects.all().aggregate(Avg('nilai_suhu'))
    max_temp = TabelSuhu.objects.all().aggregate(Max('nilai_suhu'))

    min_press = TabelTekananUdara.objects.all().aggregate(Min('nilai_tekananudara'))
    avg_press = TabelTekananUdara.objects.all().aggregate(Avg('nilai_tekananudara'))
    max_press = TabelTekananUdara.objects.all().aggregate(Max('nilai_tekananudara'))

    listing = []

    listing.append({
            'temp_time'     : datatemp.time,
            'temp_value' 	: datatemp.nilai_suhu,
            'temp_min' 	    : min_temp,
            'temp_avg' 	    : avg_temp,
            'temp_max' 	    : max_temp,
            'hum_time'      : datahum.time,
            'hum_value' 	: datahum.nilai_kelembaban,
            'hum_min' 	    : min_hum,
            'hum_avg' 	    : avg_hum,
            'hum_max' 	    : max_hum,
            'press_time'    : datapress.time,
            'press_value' 	: datapress.nilai_tekananudara,
            'press_min' 	: min_press,
            'press_avg' 	: avg_press,
            'press_max' 	: max_press
        })
    return JsonResponse(listing, safe=False)

def api_suhu(request):
    querytemp = TabelSuhu.objects.filter(api_key='DEFISR98')
    datatemp = querytemp.last()
    min_val = TabelSuhu.objects.all().aggregate(Min('nilai_suhu'))
    avg_val = TabelSuhu.objects.all().aggregate(Avg('nilai_suhu'))
    max_val = TabelSuhu.objects.all().aggregate(Max('nilai_suhu'))

    listing = []

    listing.append({
            'temp_time'     : datatemp.time,
            'temp_value' 	: datatemp.nilai_suhu,
            'temp_min' 	    : min_val,
            'temp_avg' 	    : avg_val,
            'temp_max' 	    : max_val
        })
    return JsonResponse(listing, safe=False)

def api_kelembaban(request):
    queryhum = TabelKelembaban.objects.filter(api_key='DEFISR98')
    datahum = queryhum.last()
    min_val = TabelKelembaban.objects.all().aggregate(Min('nilai_kelembaban'))
    avg_val = TabelKelembaban.objects.all().aggregate(Avg('nilai_kelembaban'))
    max_val = TabelKelembaban.objects.all().aggregate(Max('nilai_kelembaban'))

    listing = []

    listing.append({
            'hum_time'      : datahum.time,
            'hum_value' 	: datahum.nilai_kelembaban,
            'hum_min' 	    : min_val,
            'hum_avg' 	    : avg_val,
            'hum_max' 	    : max_val
        })
    return JsonResponse(listing, safe=False)

def api_tekananudara(request):
    querypress = TabelTekananUdara.objects.filter(api_key='DEFISR98')
    datapress = querypress.last()
    min_val = TabelTekananUdara.objects.all().aggregate(Min('nilai_tekananudara'))
    avg_val = TabelTekananUdara.objects.all().aggregate(Avg('nilai_tekananudara'))
    max_val = TabelTekananUdara.objects.all().aggregate(Max('nilai_tekananudara'))

    listing = []

    listing.append({
            'press_time'     : datapress.time,
            'press_value' 	 : datapress.nilai_tekananudara,
            'press_min' 	 : min_val,
            'press_avg' 	 : avg_val,
            'press_max' 	 : max_val
        })
    return JsonResponse(listing, safe=False)

@csrf_exempt
def data_cuaca(request, api_key, temp_value, hum_value, press_value):
	suhu = request.POST.get("temp_value", temp_value)
	kelembaban = request.POST.get("hum_value", hum_value)
	tekananudara = request.POST.get("press_value", press_value)

	data_suhu = TabelSuhu.objects.create(
        api_key = api_key,
        nilai_suhu  = suhu
    )
	data_suhu.save()

	data_kelembaban = TabelKelembaban.objects.create(
        api_key = api_key,
        nilai_kelembaban  = kelembaban
    )
	data_kelembaban.save()

	data_tekananudara = TabelTekananUdara.objects.create(
        api_key = api_key,
        nilai_tekananudara  = tekananudara
    )
	data_tekananudara.save()

	querytemp = TabelSuhu.objects.filter(api_key='DEFISR98')
	queryhum = TabelKelembaban.objects.filter(api_key='DEFISR98')
	querypress = TabelTekananUdara.objects.filter(api_key='DEFISR98')

	datatemp = querytemp.last()
	datahum = queryhum.last()
	datapress = querypress.last()
	mydata = []
	mydata.append({
	        'temp_time'		: datatemp.time,
	        'temp_value' 	: datatemp.nilai_suhu,
	        'hum_time'		: datahum.time,
	        'hum_value'		: datahum.nilai_kelembaban,
	        'press_time'	: datapress.time,
	        'press_value'	: datapress.nilai_tekananudara
	    })
	return JsonResponse(mydata, safe=False)